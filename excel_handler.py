"""
excel_handler.py - Excel reading and writing with openpyxl + pandas.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from models import ClassificationResult, CompanyRecord

logger = logging.getLogger(__name__)

# ── Output column order ───────────────────────────────────────────────────────
OUTPUT_COLUMNS: list[str] = [
    "S.NO",
    "CIN",
    "Company Name",
    "Domain",
    "Confidence",
    "Primary Domain Area",
    "Hardware or Software",
    "Hiring Possible",
    "Fresher Friendly",
    "Likely Roles",
    "Reason",
]

# Header colour: dark slate-blue
HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)

# Domain colour map
DOMAIN_COLOURS: dict[str, str] = {
    "ECE":     "D6E4F0",  # light blue
    "CSE":     "D5F5E3",  # light green
    "BOTH":    "FEF9E7",  # light yellow
    "NEITHER": "FADBD8",  # light red
}


# ─────────────────────────────────────────────────────────────────────────────
# Reading
# ─────────────────────────────────────────────────────────────────────────────

def _find_header_row(path: Path) -> int:
    """
    Scan the first 20 rows of an Excel file to find the line that contains
    the real column headers (the row that has 'S.NO' or 'SNO' in any cell).

    Returns the 0-based row index to pass as ``header=`` to ``pd.read_excel``.
    Defaults to 0 if no match is found.
    """
    raw = pd.read_excel(path, engine="openpyxl", header=None, nrows=20)
    for idx, row in raw.iterrows():
        values = [str(v).strip().upper().replace(".", "").replace(" ", "") for v in row]
        if "SNO" in values or "S.NO" in [str(v).strip().upper() for v in row]:
            logger.debug("Header row detected at index %d.", idx)
            return int(idx)
    logger.warning("Could not auto-detect header row; defaulting to row 0.")
    return 0


def _normalise_columns(df: pd.DataFrame) -> dict[str, str]:
    """
    Build a case-insensitive mapping from canonical names to actual df column names.

    Handles variations like 'Cin' vs 'CIN' vs 'cin', 'S.no' vs 'S.NO', etc.

    Returns:
        Dict of canonical_upper → actual_column_name.
    """
    mapping: dict[str, str] = {}
    canonical = {
        "S.NO": ["S.NO", "SNO", "S NO", "SR NO", "SR.NO", "SRNO"],
        "CIN":  ["CIN", "CIN NO", "COMPANY CIN"],
        "Company Name": ["COMPANY NAME", "NAME", "COMPANYNAME"],
    }
    normalised_cols = {
        str(c).strip().upper().replace(" ", "").replace(".", ""): str(c).strip()
        for c in df.columns
    }
    for canon, variants in canonical.items():
        for v in variants:
            key = v.replace(" ", "").replace(".", "")
            if key in normalised_cols:
                mapping[canon] = normalised_cols[key]
                break
    return mapping


def load_companies(excel_path: str) -> list[CompanyRecord]:
    """
    Read the input Excel file and return a list of CompanyRecord objects.

    Automatically handles files that have title/banner rows before the real
    column headers, and does case-insensitive column name matching so
    ``Cin``, ``CIN``, ``cin`` etc. all work correctly.

    Expected columns (any capitalisation): S.NO, CIN, Company Name
    Extra columns are silently ignored.

    Args:
        excel_path: Absolute or relative path to the .xlsx file.

    Returns:
        List of CompanyRecord instances.

    Raises:
        FileNotFoundError: If the file does not exist.
        ValueError:        If required columns are missing.
    """
    path = Path(excel_path)
    if not path.exists():
        raise FileNotFoundError(f"Input file not found: {path.resolve()}")

    logger.info("Loading companies from: %s", path.resolve())

    header_row = _find_header_row(path)
    df = pd.read_excel(path, engine="openpyxl", header=header_row)
    df.columns = [str(c).strip() for c in df.columns]

    col_map = _normalise_columns(df)
    missing = {k for k in ("S.NO", "CIN", "Company Name") if k not in col_map}
    if missing:
        raise ValueError(
            f"Missing required columns: {missing}. "
            f"Found columns: {list(df.columns)}"
        )

    sno_col  = col_map["S.NO"]
    cin_col  = col_map["CIN"]
    name_col = col_map["Company Name"]

    records: list[CompanyRecord] = []
    for _, row in df.iterrows():
        name = str(row[name_col]).strip()
        # Skip blank / NaN rows
        if not name or name.lower() in ("nan", ""):
            continue
        records.append(
            CompanyRecord(
                sno=int(row[sno_col]) if pd.notna(row[sno_col]) else len(records) + 1,
                cin=str(row[cin_col]).strip() if pd.notna(row[cin_col]) else "UNKNOWN",
                company_name=name,
            )
        )

    logger.info("Loaded %d companies (header at row %d).", len(records), header_row)
    return records


# ─────────────────────────────────────────────────────────────────────────────
# Writing
# ─────────────────────────────────────────────────────────────────────────────

def results_to_dataframe(
    records: list[CompanyRecord],
    results: dict[str, ClassificationResult],
) -> pd.DataFrame:
    """
    Merge company records with their classification results into a DataFrame.

    Args:
        records: List of input company records.
        results: Mapping of company_name → ClassificationResult.

    Returns:
        DataFrame with OUTPUT_COLUMNS columns.
    """
    rows: list[dict[str, Any]] = []
    for rec in records:
        res = results.get(rec.company_name)
        if res is None:
            rows.append(
                {
                    "S.NO": rec.sno,
                    "CIN": rec.cin,
                    "Company Name": rec.company_name,
                    "Domain": "NEITHER",
                    "Confidence": "LOW",
                    "Primary Domain Area": "UNKNOWN",
                    "Hardware or Software": "Neither",
                    "Hiring Possible": "UNKNOWN",
                    "Fresher Friendly": "UNKNOWN",
                    "Likely Roles": "UNKNOWN",
                    "Reason": "No result returned.",
                }
            )
        else:
            rows.append(
                {
                    "S.NO": rec.sno,
                    "CIN": rec.cin,
                    "Company Name": rec.company_name,
                    "Domain": res.domain,
                    "Confidence": res.confidence,
                    "Primary Domain Area": res.primary_domain_area,
                    "Hardware or Software": res.hardware_or_software,
                    "Hiring Possible": res.hiring_possible,
                    "Fresher Friendly": res.fresher_friendly,
                    "Likely Roles": res.likely_roles,
                    "Reason": res.reason,
                }
            )
    return pd.DataFrame(rows, columns=OUTPUT_COLUMNS)


def save_results(
    records: list[CompanyRecord],
    results: dict[str, ClassificationResult],
    output_path: str,
    sheet_name: str = "classified_companies",
) -> None:
    """
    Save classification results to a formatted Excel file.

    Args:
        records:     Original company records (determines row order).
        results:     Classification results keyed by company name.
        output_path: Destination .xlsx path.
        sheet_name:  Target sheet name.
    """
    df = results_to_dataframe(records, results)
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    logger.info("Saving %d rows to: %s", len(df), path.resolve())
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        _apply_formatting(writer.book[sheet_name], df)

    logger.info("Output saved successfully → %s", output_path)


def _apply_formatting(ws: Any, df: pd.DataFrame) -> None:
    """Apply header styles, column widths, and domain-based row colours."""
    # Header row
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Auto-fit column widths (cap at 60)
    for col_idx, col_name in enumerate(df.columns, start=1):
        max_len = max(
            len(str(col_name)),
            df[col_name].astype(str).str.len().max() if not df.empty else 0,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)

    # Row colouring based on Domain
    domain_col_idx = OUTPUT_COLUMNS.index("Domain") + 1
    for row_idx in range(2, ws.max_row + 1):
        domain_val = ws.cell(row=row_idx, column=domain_col_idx).value or "NEITHER"
        colour = DOMAIN_COLOURS.get(str(domain_val).upper(), "FFFFFF")
        fill = PatternFill(start_color=colour, end_color=colour, fill_type="solid")
        for cell in ws[row_idx]:
            cell.fill = fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
